import openpyxl
import pytest

XLSX = "apps/cmdb/support-files/model_config.xlsx"
DEVICE_SHEETS = [
    "attr-switch", "attr-router", "attr-firewall",
    "attr-loadbalance", "attr-physcial_server",
]


def _attr_ids(sheet):
    ids = set()
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row and row[0]:
            ids.add(str(row[0]).strip())
    return ids


@pytest.mark.unit
def test_rack_has_row_and_col():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ids = _attr_ids(wb["attr-rack"])
    assert "row" in ids
    assert "col" in ids


@pytest.mark.unit
def test_device_models_have_u_fields():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    for name in DEVICE_SHEETS:
        ids = _attr_ids(wb[name])
        assert "rack_u_start" in ids, name
        assert "u_size" in ids, name
