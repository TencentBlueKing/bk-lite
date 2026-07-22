"""CMDB ModelManage 自动关联规则导入辅助 + replace 覆盖测试。

对照 specs/capabilities/legacy-prd-cmdb-模型管理.md：模型配置导入中自动关联规则 sheet 解析、
_build_model_asst_id / _parse_auto_relation_rule_set_cell 等辅助、
replace_model_auto_relation_rule_set 全量替换。
"""

import json

import pytest

from apps.cmdb.services.model import ModelManage
from apps.core.exceptions.base_app_exception import BaseAppException

MODULE = "apps.cmdb.services.model"
_STR_ATTRS = [{"attr_id": "ip", "attr_type": "str"}, {"attr_id": "name", "attr_type": "str"}]


# --------------------------------------------------------------------------
# _build_model_asst_id
# --------------------------------------------------------------------------


def test_build_model_asst_id_ok():
    assert ModelManage._build_model_asst_id("host", "conn", "sw") == "host_conn_sw"


def test_build_model_asst_id_strips():
    assert ModelManage._build_model_asst_id("  host  ", " conn ", "sw") == "host_conn_sw"


def test_build_model_asst_id_empty_raises():
    with pytest.raises(BaseAppException):
        ModelManage._build_model_asst_id("", "conn", "sw")
    with pytest.raises(BaseAppException):
        ModelManage._build_model_asst_id("host", "", "sw")
    with pytest.raises(BaseAppException):
        ModelManage._build_model_asst_id("host", "conn", "")


# --------------------------------------------------------------------------
# _parse_auto_relation_rule_set_cell
# --------------------------------------------------------------------------


def test_parse_auto_rule_cell_dict():
    payload = {"rules": [{"match_pairs": [{"src_field_id": "a", "dst_field_id": "b"}]}]}
    out = ModelManage._parse_auto_relation_rule_set_cell(payload, context="ctx")
    assert "rules" in out


def test_parse_auto_rule_cell_json_string():
    raw = json.dumps({"rules": [{"match_pairs": [{"src_field_id": "a", "dst_field_id": "b"}]}]})
    out = ModelManage._parse_auto_relation_rule_set_cell(raw, context="ctx")
    assert "rules" in out


def test_parse_auto_rule_cell_bad_json():
    with pytest.raises(BaseAppException):
        ModelManage._parse_auto_relation_rule_set_cell("{badjson", context="ctx")


def test_parse_auto_rule_cell_empty_rules():
    with pytest.raises(BaseAppException):
        ModelManage._parse_auto_relation_rule_set_cell({"rules": []}, context="ctx")


def test_parse_auto_rule_cell_wrong_type():
    with pytest.raises(BaseAppException):
        ModelManage._parse_auto_relation_rule_set_cell(123, context="ctx")


# --------------------------------------------------------------------------
# _validate_auto_rule_sheet_authority
# --------------------------------------------------------------------------


def test_validate_sheet_authority_ok():
    ModelManage._validate_auto_rule_sheet_authority("asso-host", "host", "ctx")


def test_validate_sheet_authority_wrong():
    with pytest.raises(BaseAppException):
        ModelManage._validate_auto_rule_sheet_authority("asso-sw", "host", "ctx")


# --------------------------------------------------------------------------
# _is_empty_auto_rule_sheet_row
# --------------------------------------------------------------------------


def test_is_empty_auto_rule_sheet_row_empty():
    assert ModelManage._is_empty_auto_rule_sheet_row({}) is True
    assert ModelManage._is_empty_auto_rule_sheet_row(
        {"src_model_id": " ", "dst_model_id": "", "asst_id": None, "auto_relation_rule": ""}
    ) is True


def test_is_empty_auto_rule_sheet_row_nonempty():
    assert ModelManage._is_empty_auto_rule_sheet_row(
        {"src_model_id": "host", "dst_model_id": "sw", "asst_id": "conn", "auto_relation_rule": ""}
    ) is False


# --------------------------------------------------------------------------
# _import_auto_relation_rule_sets_from_asso_sheets
# --------------------------------------------------------------------------


@pytest.mark.django_db
def test_import_auto_rule_sets_empty():
    # 空 model_config → 不抛
    ModelManage._import_auto_relation_rule_sets_from_asso_sheets({})


@pytest.mark.django_db
def test_import_auto_rule_sets_skips_non_asso():
    ModelManage._import_auto_relation_rule_sets_from_asso_sheets({"model": [], "attr-host": []})


@pytest.mark.django_db
def test_import_auto_rule_sets_wrong_sheet_raises():
    """auto_relation_rule 必须定义在 asso-<src_model_id>，否则抛"""
    payload = {"rules": [{"match_pairs": [{"src_field_id": "ip", "dst_field_id": "ip"}]}]}
    rows = [
        {
            "src_model_id": "host", "dst_model_id": "sw", "asst_id": "conn",
            "auto_relation_rule": json.dumps(payload),
        }
    ]
    with pytest.raises(BaseAppException):
        # sheet 名是 asso-sw，但 src 是 host → 不匹配
        ModelManage._import_auto_relation_rule_sets_from_asso_sheets({"asso-sw": rows})


@pytest.mark.django_db
def test_import_auto_rule_sets_duplicate_raises():
    payload = {"rules": [{"match_pairs": [{"src_field_id": "ip", "dst_field_id": "ip"}]}]}
    rows = [
        {"src_model_id": "host", "dst_model_id": "sw", "asst_id": "conn",
         "auto_relation_rule": json.dumps(payload)},
        {"src_model_id": "host", "dst_model_id": "sw", "asst_id": "conn",
         "auto_relation_rule": json.dumps(payload)},
    ]
    with pytest.raises(BaseAppException):
        ModelManage._import_auto_relation_rule_sets_from_asso_sheets({"asso-host": rows})


@pytest.mark.django_db
def test_import_auto_rule_sets_ok(fake_graph, monkeypatch):
    monkeypatch.setattr(f"{MODULE}.create_change_record", lambda **k: None)
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_info_search",
        lambda mid: {"_id": 1, "src_model_id": "host", "dst_model_id": "sw"},
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage._get_model_attrs_for_auto_rule",
        lambda mid: _STR_ATTRS,
    )
    monkeypatch.setattr(
        "apps.cmdb.services.auto_relation_reconcile.schedule_rule_auto_relation_full_sync",
        lambda ids: None,
    )
    fg = fake_graph(MODULE)
    payload = {"rules": [{"match_pairs": [{"src_field_id": "ip", "dst_field_id": "ip", "matching_rule": "exact"}]}]}
    rows = [
        {"src_model_id": "host", "dst_model_id": "sw", "asst_id": "conn",
         "auto_relation_rule": json.dumps(payload)},
    ]
    ModelManage._import_auto_relation_rule_sets_from_asso_sheets({"asso-host": rows})
    # 导入成功后应触发 set_edge_properties 写回规则
    assert any(c[0] == "set_edge_properties" for c in fg.calls)


# --------------------------------------------------------------------------
# replace_model_auto_relation_rule_set
# --------------------------------------------------------------------------


@pytest.mark.django_db
def test_replace_auto_relation_no_assoc(monkeypatch):
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_info_search", lambda mid: {}
    )
    with pytest.raises(BaseAppException):
        ModelManage.replace_model_auto_relation_rule_set("host", "x", {"rules": []})


@pytest.mark.django_db
def test_replace_auto_relation_wrong_model(monkeypatch):
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_info_search",
        lambda mid: {"_id": 1, "src_model_id": "a", "dst_model_id": "b"},
    )
    with pytest.raises(BaseAppException):
        ModelManage.replace_model_auto_relation_rule_set("nope", "x", {"rules": []})


# --------------------------------------------------------------------------
# export_model_config（端到端 mock）
# --------------------------------------------------------------------------


@pytest.mark.django_db
def test_export_model_config_minimal(monkeypatch):
    monkeypatch.setattr(
        f"{MODULE}.ClassificationManage.search_model_classification",
        lambda language="en": [{"classification_id": "net", "classification_name": "网络"}],
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.search_model",
        lambda language="en": [{"model_id": "host", "model_name": "主机",
                                "icn": "icon", "classification_id": "net", "attrs": "[]"}],
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_search", lambda mid: [],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.public_enum_library.list_libraries", lambda: [],
    )
    # build_unique_rule_context 内部用 search_model_info
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.search_model_info",
        lambda mid: {"model_id": mid, "attrs": "[]", "unique_rules": "[]", "_id": 1},
    )
    stream = ModelManage.export_model_config(language="zh-Hans")
    data = stream.read()
    assert data[:2] == b"PK"


@pytest.mark.django_db
def test_export_model_config_with_attrs_and_assoc(monkeypatch):
    attrs = [
        {"attr_id": "inst_name", "attr_name": "名称", "attr_type": "str",
         "is_required": True, "editable": True, "is_only": True},
        {"attr_id": "status", "attr_name": "状态", "attr_type": "enum",
         "enum_rule_type": "custom", "option": [{"id": "1", "name": "运行"}]},
    ]
    monkeypatch.setattr(
        f"{MODULE}.ClassificationManage.search_model_classification",
        lambda language="en": [],
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.search_model",
        lambda language="en": [{"model_id": "host", "attrs": json.dumps(attrs),
                                "model_name": "主机", "icn": "icon", "classification_id": "net"}],
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_search",
        lambda mid: [{"src_model_id": "host", "dst_model_id": "sw", "asst_id": "conn",
                      "mapping": "1:n"}],
    )
    monkeypatch.setattr(
        "apps.cmdb.services.public_enum_library.list_libraries",
        lambda: [{"library_id": "lib_1", "name": "状态库", "team": [1],
                  "options": [{"id": "1", "name": "运行"}]}],
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.search_model_info",
        lambda mid: {"model_id": mid, "attrs": json.dumps(attrs), "unique_rules": "[]", "_id": 1},
    )
    stream = ModelManage.export_model_config(language="en")
    assert stream.read()[:2] == b"PK"


@pytest.mark.django_db
def test_replace_auto_relation_ok(fake_graph, monkeypatch):
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_info_search",
        lambda mid: {"_id": 1, "src_model_id": "host", "dst_model_id": "sw"},
    )
    monkeypatch.setattr(
        f"{MODULE}.ModelManage._get_model_attrs_for_auto_rule", lambda mid: _STR_ATTRS
    )
    monkeypatch.setattr(
        "apps.cmdb.services.auto_relation_reconcile.schedule_rule_auto_relation_full_sync",
        lambda ids: None,
    )
    fake_graph(MODULE)
    payload = {
        "rules": [
            {"rule_id": "r1",
             "match_pairs": [{"src_field_id": "ip", "dst_field_id": "ip", "matching_rule": "exact"}]}
        ]
    }
    result = ModelManage.replace_model_auto_relation_rule_set("host", "x", payload, username="admin")
    assert len(result.rules) == 1


# --------------------------------------------------------------------------
# export_model_config 勾选过滤
# --------------------------------------------------------------------------


def _read_export(stream):
    """把导出的 xlsx 字节流读回 {sheet_title: [rows...]}。"""
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(stream.read()))
    return {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb.worksheets}


def _setup_two_models(monkeypatch):
    """两个模型 host(net 分类) / sw(net2 分类)，host->sw 关联一条。"""
    classifications = [
        {"classification_id": "net", "classification_name": "网络"},
        {"classification_id": "net2", "classification_name": "网络2"},
    ]
    models = [
        {"model_id": "host", "model_name": "主机", "icn": "i1",
         "classification_id": "net", "attrs": "[]"},
        {"model_id": "sw", "model_name": "交换机", "icn": "i2",
         "classification_id": "net2", "attrs": "[]"},
    ]
    monkeypatch.setattr(
        f"{MODULE}.ClassificationManage.search_model_classification",
        lambda language="en": classifications,
    )
    monkeypatch.setattr(f"{MODULE}.ModelManage.search_model", lambda language="en": models)
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.model_association_search",
        lambda mid: [{"src_model_id": "host", "dst_model_id": "sw",
                      "asst_id": "conn", "mapping": "1:n"}] if mid == "host" else [],
    )
    monkeypatch.setattr("apps.cmdb.services.public_enum_library.list_libraries",
                        lambda: [{"library_id": "lib_1", "name": "库", "team": [1], "options": []}])
    monkeypatch.setattr(
        f"{MODULE}.ModelManage.search_model_info",
        lambda mid: {"model_id": mid, "attrs": "[]", "unique_rules": "[]", "_id": 1},
    )


@pytest.mark.django_db
def test_export_model_config_subset_filters_models_and_classifications(monkeypatch):
    _setup_two_models(monkeypatch)
    sheets = _read_export(ModelManage.export_model_config(language="zh-Hans", model_ids=["host"]))

    model_data_rows = sheets["models"][2:]
    assert [r[0] for r in model_data_rows] == ["host"]
    assert "attr-host" in sheets and "attr-sw" not in sheets
    cls_data_rows = sheets["classifications"][2:]
    assert [r[0] for r in cls_data_rows] == ["net"]
    assert sheets["public_enum_libraries"][2:][0][0] == "lib_1"


@pytest.mark.django_db
def test_export_model_config_subset_drops_cross_selection_assoc(monkeypatch):
    _setup_two_models(monkeypatch)
    sheets = _read_export(ModelManage.export_model_config(language="zh-Hans", model_ids=["host"]))
    assert "asso-host" not in sheets


@pytest.mark.django_db
def test_export_model_config_keeps_assoc_when_both_selected(monkeypatch):
    _setup_two_models(monkeypatch)
    sheets = _read_export(ModelManage.export_model_config(language="zh-Hans", model_ids=["host", "sw"]))
    assert "asso-host" in sheets
    asso_rows = sheets["asso-host"][2:]
    assert asso_rows[0][0] == "host" and asso_rows[0][1] == "sw"


@pytest.mark.django_db
def test_export_model_config_empty_model_ids_exports_all(monkeypatch):
    _setup_two_models(monkeypatch)
    sheets = _read_export(ModelManage.export_model_config(language="zh-Hans", model_ids=[]))
    assert [r[0] for r in sheets["models"][2:]] == ["host", "sw"]
    assert {r[0] for r in sheets["classifications"][2:]} == {"net", "net2"}
