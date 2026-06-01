"""CMDB Export 工具纯逻辑覆盖测试（无 DB；Export 不传 association 时不调图库）。

对照 spec/prd/CMDB·实例导出：Tag 序列化、用户显示名格式化、Excel 模板表头/枚举校验/导出
inst_list 与模板生成。
"""

import openpyxl
import pytest

from apps.cmdb.utils.export import Export, serialize_tag_values_for_export


# --------------------------------------------------------------------------
# serialize_tag_values_for_export
# --------------------------------------------------------------------------


def test_serialize_tag_values_empty():
    assert serialize_tag_values_for_export([]) == ""
    assert serialize_tag_values_for_export(None) == ""


def test_serialize_tag_values_filters_blank():
    assert serialize_tag_values_for_export(["env:prod", "  ", "app:web"]) == "env:prod,app:web"


def test_serialize_tag_values_strips():
    assert serialize_tag_values_for_export([" a:1 "]) == "a:1"


# --------------------------------------------------------------------------
# _format_user_display_username
# --------------------------------------------------------------------------


def test_format_user_display_none():
    assert Export._format_user_display_username(None) is None
    assert Export._format_user_display_username({}) is None


def test_format_user_display_full():
    out = Export._format_user_display_username({"username": "alice", "display_name": "张三"})
    assert out == "张三(alice)"


def test_format_user_display_name_only():
    assert Export._format_user_display_username({"display_name": "张三"}) == "张三"


def test_format_user_display_username_only():
    assert Export._format_user_display_username({"username": "alice"}) == "alice"


def test_format_user_display_name_fallback():
    # name 字段作为 username 后备
    assert Export._format_user_display_username({"name": "alice", "display_name": "张三"}) == "张三(alice)"


# --------------------------------------------------------------------------
# generate_header / export_template / return_bytesio
# --------------------------------------------------------------------------


_ATTRS = [
    {"attr_id": "inst_name", "attr_name": "实例名", "attr_type": "str", "is_required": True},
    {"attr_id": "ip", "attr_name": "IP", "attr_type": "str"},
    {"attr_id": "status", "attr_name": "状态", "attr_type": "enum", "option": [{"id": "1", "name": "运行"}, {"id": "2", "name": "停止"}]},
    {"attr_id": "display_only", "attr_name": "display", "attr_type": "str", "is_display_field": True},
]


def test_generate_header_basics():
    wb = Export(_ATTRS, model_id="host").generate_header()
    sheet = wb.active
    assert sheet.title == "host"
    # 第 1/2/3 行：字段名、字段类型、字段标识；display_field 应被过滤
    names = [c.value for c in sheet[1]]
    ids = [c.value for c in sheet[3]]
    assert "实例名(必填)" in names
    assert "IP" in names
    assert "display_only" not in ids


def test_generate_header_enum_creates_validation_sheet():
    wb = Export(_ATTRS, model_id="host").generate_header()
    # 应为 enum 字段额外创建以字段名命名的 sheet
    assert "状态" in wb.sheetnames


def test_export_template_returns_bytesio():
    stream = Export(_ATTRS, model_id="host").export_template()
    data = stream.read()
    assert data[:2] == b"PK"  # xlsx zip 头


# --------------------------------------------------------------------------
# set_row_color / set_cell_color
# --------------------------------------------------------------------------


def test_set_row_and_cell_color():
    obj = Export(_ATTRS, model_id="host")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["a", "b", "c"])
    obj.set_row_color(sheet, 1, "FF0000")
    obj.set_cell_color(sheet, 1, 1, "00FF00")
    # 单元格应有 fill 属性（具体 PatternFill 已设置）
    fill = sheet.cell(row=1, column=1).fill
    assert fill.fill_type == "solid"


# --------------------------------------------------------------------------
# export_inst_list（无关联，避免触图）
# --------------------------------------------------------------------------


def test_export_inst_list_basic():
    inst_list = [
        {"_id": 1, "inst_name": "h1", "ip": "1.1.1.1", "status": ["1"]},
        {"_id": 2, "inst_name": "h2", "ip": "2.2.2.2"},
    ]
    # 不传 association → format_inst_asst_name 跳过关联查询
    stream = Export(_ATTRS, model_id="host").export_inst_list(inst_list)
    data = stream.read()
    assert data[:2] == b"PK"


def test_export_inst_list_empty():
    stream = Export(_ATTRS, model_id="host").export_inst_list([])
    data = stream.read()
    assert data[:2] == b"PK"
