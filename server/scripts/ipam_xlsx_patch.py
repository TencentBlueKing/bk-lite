"""Idempotent patch: add IPAM increments to model_config.xlsx.

Increments (mirrors ipam_init.py):
  attr-subnet: gateway(str), vlan_id(int), usage_type(enum), owner(user)
  attr-ip:     description(str)
  asso-ip:     ip-->use-->host (n:n), ip-->use-->network (n:n)

Safe to run multiple times — each entry is checked by attr_id / (src,asst,dst)
before appending.
"""

import json
import os
import sys

import openpyxl

XLSX_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "apps",
    "cmdb",
    "support-files",
    "model_config.xlsx",
)
XLSX_PATH = os.path.normpath(XLSX_PATH)

# ---------------------------------------------------------------------------
# New subnet attrs — column order matches attr-subnet sheet (9 cols):
#   attr_id | attr_name | attr_type | option | attr_group |
#   is_only | editable  | is_required | user_prompt
# ---------------------------------------------------------------------------
SUBNET_NEW_ATTRS = [
    (
        "gateway",
        "网关",
        "str",
        '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
        "基本信息",
        False,
        True,
        False,
        None,
    ),
    (
        "vlan_id",
        "VLAN",
        "int",
        '{"min_value": "", "max_value": ""}',
        "基本信息",
        False,
        True,
        False,
        None,
    ),
    (
        "usage_type",
        "用途类型",
        "enum",
        json.dumps(
            [
                {"id": "business", "name": "业务"},
                {"id": "management", "name": "管理"},
                {"id": "dmz", "name": "DMZ"},
                {"id": "interconnect", "name": "互联"},
                {"id": "other", "name": "其他"},
            ],
            ensure_ascii=False,
        ),
        "基本信息",
        False,
        True,
        False,
        None,
    ),
    (
        "owner",
        "负责人",
        "user",
        None,
        "基本信息",
        False,
        True,
        False,
        None,
    ),
]

# ---------------------------------------------------------------------------
# New ip attrs — column order matches attr-ip sheet (10 cols):
#   attr_id | attr_name | attr_type | option | attr_group |
#   is_only | editable  | is_required | <None col> | user_prompt
# ---------------------------------------------------------------------------
IP_NEW_ATTRS = [
    (
        "description",
        "描述",
        "str",
        '{"validation_type":"unrestricted","custom_regex":"","widget_type":"single_line"}',
        "基本信息",
        False,
        True,
        False,
        None,
        None,
    ),
]

# ---------------------------------------------------------------------------
# New ip associations — column order matches asso-subnet / asso-k8s_namespace:
#   src_model_id | dst_model_id | asst_id | mapping
# (asso-subnet has 9 cols but only 4 used; we write only the 4 data cols)
# ---------------------------------------------------------------------------
IP_NEW_ASSOS = [
    ("ip", "host", "use", "n:n"),
    ("ip", "network", "use", "n:n"),
]

ASSO_IP_HEADERS = (
    ("源模型", "目标模型", "关联关系", "源-目标约束"),
    ("src_model_id", "dst_model_id", "asst_id", "mapping"),
)


def _get_existing_attr_ids(ws, header_row=2):
    """Collect attr_id values from rows > header_row (col A = attr_id)."""
    ids = set()
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=1, values_only=True):
        val = row[0]
        if val:
            ids.add(str(val).strip())
    return ids


def _get_existing_assos(ws, header_row=2):
    """Return set of (src_model_id, asst_id, dst_model_id) tuples."""
    assos = set()
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=3, values_only=True):
        src, dst, asst = row[0], row[1], row[2]
        if src and dst and asst:
            assos.add((str(src).strip(), str(asst).strip(), str(dst).strip()))
    return assos


def _find_next_empty_row(ws, col=1):
    """Find the first row (>= 3) where column `col` is empty."""
    for row_idx in range(3, ws.max_row + 2):
        val = ws.cell(row=row_idx, column=col).value
        if not val:
            return row_idx
    return ws.max_row + 1


def patch_xlsx(path):
    print(f"Loading workbook: {path}")
    wb = openpyxl.load_workbook(path)

    added = []
    skipped = []

    # ------------------------------------------------------------------
    # 1. attr-subnet
    # ------------------------------------------------------------------
    ws_subnet = wb["attr-subnet"]
    existing_ids = _get_existing_attr_ids(ws_subnet)
    for row_data in SUBNET_NEW_ATTRS:
        attr_id = row_data[0]
        if attr_id in existing_ids:
            skipped.append(f"attr-subnet/{attr_id}")
            continue
        next_row = _find_next_empty_row(ws_subnet)
        for col_idx, val in enumerate(row_data, start=1):
            ws_subnet.cell(row=next_row, column=col_idx, value=val)
        added.append(f"attr-subnet/{attr_id}")
        existing_ids.add(attr_id)

    # ------------------------------------------------------------------
    # 2. attr-ip
    # ------------------------------------------------------------------
    ws_ip = wb["attr-ip"]
    existing_ids = _get_existing_attr_ids(ws_ip)
    for row_data in IP_NEW_ATTRS:
        attr_id = row_data[0]
        if attr_id in existing_ids:
            skipped.append(f"attr-ip/{attr_id}")
            continue
        next_row = _find_next_empty_row(ws_ip)
        for col_idx, val in enumerate(row_data, start=1):
            ws_ip.cell(row=next_row, column=col_idx, value=val)
        added.append(f"attr-ip/{attr_id}")
        existing_ids.add(attr_id)

    # ------------------------------------------------------------------
    # 3. asso-ip (create sheet if missing, then add association rows)
    # ------------------------------------------------------------------
    if "asso-ip" not in wb.sheetnames:
        ws_asso_ip = wb.create_sheet("asso-ip")
        # Write header rows (row 1 = labels, row 2 = field names)
        for r_idx, hdr_row in enumerate(ASSO_IP_HEADERS, start=1):
            for c_idx, val in enumerate(hdr_row, start=1):
                ws_asso_ip.cell(row=r_idx, column=c_idx, value=val)
        print("Created new sheet: asso-ip")
        existing_assos = set()
    else:
        ws_asso_ip = wb["asso-ip"]
        existing_assos = _get_existing_assos(ws_asso_ip)

    for src, dst, asst, mapping in IP_NEW_ASSOS:
        key = (src, asst, dst)
        if key in existing_assos:
            skipped.append(f"asso-ip/{src}_{asst}_{dst}")
            continue
        next_row = _find_next_empty_row(ws_asso_ip)
        for col_idx, val in enumerate((src, dst, asst, mapping), start=1):
            ws_asso_ip.cell(row=next_row, column=col_idx, value=val)
        added.append(f"asso-ip/{src}_{asst}_{dst}")
        existing_assos.add(key)

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    wb.save(path)
    print(f"\nSaved: {path}")
    print(f"Added   ({len(added)}): {added}")
    print(f"Skipped ({len(skipped)}): {skipped}")
    return added, skipped


if __name__ == "__main__":
    added, skipped = patch_xlsx(XLSX_PATH)
    if not added and skipped:
        print("\n[IDEMPOTENT] Nothing new to add — all entries already present.")
        sys.exit(0)
    print("\n[DONE]")
